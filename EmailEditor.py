'''
this programe edits and update and excel file from
 @helpinghands.org to @helpinghands.cm
'''

import openpyxl as xl
from openpyxl.utils import get_column_letter

workbook = xl.load_workbook('transactions.xlsx')
# Define variable to read the worksheet
worksheet = workbook.active
# iterate the loop to read the cell values

for i in range(2, 5):
    #get the email column
    email_col = worksheet.cell(row=i, column=3).value
    #get the last three characters of the email
    cm_code = email_col[-4] + email_col[-3] + email_col[-2] + email_col[-1]
    #replace the email pattern
    replacement = {cm_code: ".cm"}
    for key in replacement.keys():
        if str(email_col) == key:
            newCell = replacement.get(key)
            worksheet[get_column_letter(i + 1)] = str(newCell)
            #save
workbook.save("transactions1.xlsx")

workbook1 = xl.load_workbook('transactions1.xlsx')
# Define variable to read the worksheet
worksheet1 = workbook.active

for i in range(0, worksheet1.max_row):
    for col in worksheet1.iter_cols(1, worksheet1.max_column):
        print(col[i].value, end="\t\t")
    print("")

